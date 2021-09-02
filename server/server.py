import flask
from flask import request
from flask import Response
import openpyxl
from Album import Album

priority_column = 0
artist_column = 1
album_column = 2
got_em_column = 3
price_column = 4
link_column = 5
excel_location = "D:\documenten\Albums.xlsx"
album_sheet = 'albums'

# configure API
app = flask.Flask(__name__)
app.config["DEBUG"] = True

@app.route('/gotem', methods=['POST'])
def mark_album_got_em():
    try:
        book = openpyxl.load_workbook(excel_location)
        sheet = book[album_sheet]
        data = request.get_json(True)
        album = get_album_from_request_data(data)
        print('Received album ' + album.album_name + ' by ' + album.artist)

        for row in sheet.iter_rows():
            if (row[artist_column].value == album.artist and row[album_column].value == album.album_name):
                print(album.album_name + " found. Updating Got'Em! value.")
                row[got_em_column].value = True
                book.save(excel_location)
                return {'message': 'updated album'}, 200

        print(album.album_name + " doesn't exist in the list yet. Adding it now.")
        sheet.append(album.to_tuple())
        book.save(excel_location)
        return {'message': 'created album'}, 204
    except:
        return {'error': 'Failed to access excel document. Please close the file first'}, 400


@app.route('/albums', methods=['POST'])
def add_album_to_excel():
    try:
        book = openpyxl.load_workbook(excel_location)
        sheet = book[album_sheet]
        data = request.get_json(True)
        album = get_album_from_request_data(data)
        sheet.append(album.to_tuple())
        book.save(excel_location)
        return {'message': 'created album'}, 204
    except:
        return {'error': 'Failed to access excel document. Please close the file first'}, 400



# returns a tuple of the album object
def get_album_from_request_data(data):
    priority = None
    artist = None
    album = None
    got_em = None
    price = None
    link = None

    if "priority" in data: priority = data['priority']
    if priority is None: priority = 'normal'
    if "artist" in data: artist = data['artist']
    if "album" in data: album = data['album']
    if "got_em" in data: got_em = data['got_em']
    if got_em is None: got_em = False
    if "price" in data: price = data['price']
    if "link" in data: link = data['link']

    return Album(priority, artist, album, got_em, price, link)

app.run()
